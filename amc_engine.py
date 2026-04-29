"""
amc_engine.py
Reads filled AMCFORMULA patient files and appends rows to the tracker.
No Excel, no COM, no SharePoint issues - pure file I/O via openpyxl.

Usage:
    python amc_engine.py <company_key|all> [--dry-run] [--no-archive]
"""

import sys
import os
import shutil
import json
import argparse
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import load_workbook

# ── Column letter helpers ─────────────────────────────────────────────────────

def col_letter_to_index(letter: str) -> int:
    """'A'->1, 'B'->2, 'AK'->37, etc."""
    idx = 0
    for c in letter.upper():
        idx = idx * 26 + (ord(c) - ord('A') + 1)
    return idx

def col_index_to_letter(n: int) -> str:
    result = ''
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(ord('A') + rem) + result
    return result

# ── Config ────────────────────────────────────────────────────────────────────

# Verified from actual tracker header row
TRACKER_COLUMNS = {
    'serial':     'A',
    'date_amc':   'B',
    'date_review':'C',
    'iqama':      'D',
    'name':       'E',
    'company':    'F',
    'height':     'I',
    'weight':     'J',
    'bmi':        'K',
    'age':        'AH',
    'bp':         'AI',
    'status':     'AJ',
    'comment':    'AK',
}

# Formula row -> tracker column (from AMCFORMULA Field sheet inspection)
# N=SERUM BILIRUBIN, U=HDL, X=URIC ACID are tracker-only (not in formula)
TEST_MAP = [
    (15, 'G'),   # IM CONSLTATION
    (16, 'H'),   # OPTHA. CONSULTATION
    (20, 'L'),   # CBC
    (21, 'M'),   # FBS
    (24, 'O'),   # SGPT/ALT
    (25, 'P'),   # SGOT/AST
    (26, 'Q'),   # GGT
    (27, 'R'),   # SERUM ALKALINE PHOSPHATASE
    (28, 'S'),   # SERUM CHOLE
    (29, 'T'),   # SERUM TRIGLYCERIDE
    (30, 'V'),   # LDL
    (32, 'W'),   # SERUM CREATININE
    (33, 'Y'),   # TSH
    (35, 'Z'),   # UA
    (34, 'AA'),  # STOOL
    (40, 'AB'),  # X-RAY
    (37, 'AC'),  # AUDIO
    (38, 'AD'),  # SPIRO
    (43, 'AE'),  # ECG
    (45, 'AF'),  # PSA (>=40 only)
    (42, 'AG'),  # Abdominal Ultrasound
]

# Status checkmark cells (VBA writes ✓ into I4:I7)
STATUS_CELLS = {
    'I4': 'FIT',
    'I5': 'UNFIT',
    'I6': 'CLINIC VISIT',
    'I7': 'FOR FURTHER EVALUATION',
}

COMPANIES = {
    'altamimi':    ('Al Tamimi',     'Al Tamimi'),
    'mixed':       ('Mixed',         'Mixed'),
    'alsalem':     ('AL SALEM',      'AL SALEM'),
    'sar':         ('SAR',           'SAR'),
    'jal':         ('JAL',           'JAL'),
    'alamshawi':   ('AL AMSHAWI',    'AL AMSHAWI'),
    'catrion':     ('CATRION',       'CATRION'),
    'scms':        ('SCMS',          'SCMS'),
    'alsuwaidi':   ('AL SUWAIDI',    'AL SUWAIDI'),
    'aljari':      ('AL JARI',       'AL JARI'),
    'alseif':      ('AL SEIF',       'AL SEIF'),
    'malakalreem': ('MALAK AL REEM', 'MALAK AL REEM'),
    'almutairi':   ('AL MUTAIRI',    'AL MUTAIRI'),
    'nal':         ('NAL',           'NAL'),
    'reda':        ('REDA',          'REDA'),
}

# ── Detection helpers ─────────────────────────────────────────────────────────

def is_abnormal(cell) -> bool:
    """Return True if the cell has a non-default fill colour (doctor marked it red)."""
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return False
    if fill.fill_type not in ('solid', 'gray125', 'darkGray', 'mediumGray', 'lightGray'):
        return False
    try:
        rgb = fill.fgColor.rgb
        if rgb in (None, '00000000', 'FFFFFFFF', '00FFFFFF'):
            return False
        return True
    except Exception:
        return False

def safe_str(val) -> str:
    """Convert any cell value to a clean string, handling ints/floats/dates."""
    if val is None:
        return ''
    if isinstance(val, (date, datetime)):
        return val.isoformat()
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()

def iqama_str(val) -> str:
    """Return Iqama as a clean 10-digit string."""
    if val is None:
        return ''
    try:
        return str(int(float(str(val).strip())))
    except Exception:
        return str(val).strip()

# ── Patient file reader ───────────────────────────────────────────────────────

def read_patient(path: Path) -> dict:
    wb = load_workbook(path, keep_vba=True, data_only=True)
    ws = wb['Field']

    data = {
        'name':       safe_str(ws['C4'].value),
        'company':    safe_str(ws['C5'].value),
        'iqama':      iqama_str(ws['C6'].value),
        'age':        ws['C7'].value,
        'date_amc':   ws['C8'].value,
        'date_review':ws['C9'].value,
        'bp':         safe_str(ws['C11'].value),
        'height':     ws['E11'].value,
        'weight':     ws['G11'].value,
        'comment':    safe_str(ws['B48'].value),
        'status':     None,
        'tests':      {},
    }

    # Status
    for cell_addr, label in STATUS_CELLS.items():
        val = ws[cell_addr].value
        if val is not None and str(val).strip():
            data['status'] = label
            break

    # Test results
    age_val = 0
    try:
        age_val = int(float(str(data['age']))) if data['age'] is not None else 0
    except Exception:
        pass

    for formula_row, tracker_col in TEST_MAP:
        if tracker_col == 'AF' and age_val < 40:   # PSA only >= 40
            data['tests'][tracker_col] = None
            continue
        g_cell = ws.cell(row=formula_row, column=7)
        data['tests'][tracker_col] = 'ABNORMAL' if is_abnormal(g_cell) else 'NORMAL'

    wb.close()
    return data

# ── Tracker helpers ───────────────────────────────────────────────────────────

def get_next_empty_row(ws) -> int:
    """
    A row is truly empty only when BOTH col A (SN) and col E (Name) are blank.
    This skips ghost rows that have only a serial number from a previous partial run.
    """
    row = 2
    while True:
        sn_val   = ws.cell(row=row, column=1).value
        name_val = ws.cell(row=row, column=5).value
        sn_empty   = sn_val   is None or str(sn_val).strip()   == ''
        name_empty = name_val is None or str(name_val).strip() == ''
        if sn_empty and name_empty:
            return row
        row += 1

def get_next_sn(ws, next_row: int) -> int:
    """Walk backwards from next_row to find the last real serial number."""
    for r in range(next_row - 1, 1, -1):
        sn   = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=5).value
        if sn is not None and str(sn).strip() and \
           name is not None and str(name).strip():
            try:
                return int(float(str(sn))) + 1
            except Exception:
                pass
    return 1

def iqama_exists(ws, iqama: str) -> bool:
    row = 2
    while True:
        sn = ws.cell(row=row, column=1).value
        if sn is None or str(sn).strip() == '':
            break
        existing = iqama_str(ws.cell(row=row, column=4).value)
        if existing == iqama:
            return True
        row += 1
    return False

def write_row(ws, row_idx: int, sn: int, patient: dict):
    """Write one patient into the tracker sheet at row_idx."""
    tc = TRACKER_COLUMNS

    ws.cell(row=row_idx, column=col_letter_to_index(tc['serial'])).value      = sn
    ws.cell(row=row_idx, column=col_letter_to_index(tc['date_amc'])).value    = patient['date_amc']
    ws.cell(row=row_idx, column=col_letter_to_index(tc['date_review'])).value = patient['date_review']
    ws.cell(row=row_idx, column=col_letter_to_index(tc['name'])).value        = patient['name']
    ws.cell(row=row_idx, column=col_letter_to_index(tc['company'])).value     = patient['company']
    ws.cell(row=row_idx, column=col_letter_to_index(tc['height'])).value      = patient['height']
    ws.cell(row=row_idx, column=col_letter_to_index(tc['weight'])).value      = patient['weight']
    ws.cell(row=row_idx, column=col_letter_to_index(tc['age'])).value         = patient['age']
    ws.cell(row=row_idx, column=col_letter_to_index(tc['bp'])).value          = patient['bp']
    ws.cell(row=row_idx, column=col_letter_to_index(tc['status'])).value      = patient['status']
    ws.cell(row=row_idx, column=col_letter_to_index(tc['comment'])).value     = patient['comment']

    # BMI formula
    ws.cell(row=row_idx, column=col_letter_to_index(tc['bmi'])).value = \
        f'=J{row_idx}/(I{row_idx}/100)^2'

    # Iqama as plain string (preserves all 10 digits)
    iq_cell = ws.cell(row=row_idx, column=col_letter_to_index(tc['iqama']))
    iq_cell.value = patient['iqama']
    iq_cell.number_format = '@'

    # Test results
    for tracker_col, result in patient['tests'].items():
        if result is not None:
            ws.cell(row=row_idx, column=col_letter_to_index(tracker_col)).value = result

# ── Archive ───────────────────────────────────────────────────────────────────

def archive_file(src: Path, archive_dir: Path):
    archive_dir.mkdir(parents=True, exist_ok=True)
    dest = archive_dir / src.name
    if dest.exists():
        stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        dest = archive_dir / f"{src.stem}_{stamp}{src.suffix}"
    shutil.copy2(src, dest)
    src.unlink()
    print(f"    Archived -> {dest}")

    # PDF
    pdf_src = src.with_suffix('.pdf')
    if pdf_src.exists():
        pdf_dest = archive_dir / pdf_src.name
        if pdf_dest.exists():
            stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
            pdf_dest = archive_dir / f"{pdf_src.stem}_{stamp}.pdf"
        shutil.copy2(pdf_src, pdf_dest)
        pdf_src.unlink()

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('company', nargs='?', default='all')
    parser.add_argument('--dry-run',    action='store_true')
    parser.add_argument('--no-archive', action='store_true')
    parser.add_argument('--root',       default=None)
    args = parser.parse_args()

    # Resolve root dir - passed from bat file
    if args.root:
        root = Path(args.root)
    else:
        root = Path(__file__).parent.parent

    tracker_path  = root / 'Tracker' / 'Contractors_AMC_Tracker_2026.xlsm'
    companies_dir = root / 'Companies'
    archive_dir   = root / 'Archive'
    logs_dir      = root / 'Logs'

    for d in [companies_dir, archive_dir, logs_dir]:
        d.mkdir(parents=True, exist_ok=True)

    print()
    print('  ============================================================')
    print(f"  AMC Automation  |  company={args.company}  |  dry-run={args.dry_run}")
    print('  ============================================================')
    print()

    if not tracker_path.exists():
        print(f"  ERROR: Tracker not found: {tracker_path}")
        sys.exit(1)

    # Resolve companies to process
    company_key = args.company.lower()
    if company_key == 'all':
        keys = sorted(COMPANIES.keys())
    elif company_key in COMPANIES:
        keys = [company_key]
    else:
        print(f"  ERROR: Unknown company '{args.company}'")
        print(f"  Valid: {', '.join(sorted(COMPANIES.keys()))}")
        sys.exit(1)

    # Backup tracker
    if not args.dry_run:
        logs_dir.mkdir(exist_ok=True)
        stamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        bkp = logs_dir / f'tracker-backup-{stamp}.xlsm'
        try:
            shutil.copy2(tracker_path, bkp)
            print(f"  Tracker backed up -> {bkp}")
        except Exception as e:
            print(f"  WARNING: Backup failed ({e}). Continuing.")

    # Load tracker once
    print("  Loading tracker...")
    try:
        tracker_wb = load_workbook(tracker_path, keep_vba=True)
    except Exception as e:
        print(f"  ERROR: Cannot open tracker: {e}")
        sys.exit(1)

    total_written  = 0
    total_skipped  = 0
    total_errors   = 0
    company_stats  = {}
    start_time     = datetime.now()

    for i, key in enumerate(keys, 1):
        sheet_name, folder_name = COMPANIES[key]
        folder_path = companies_dir / folder_name
        company_stats[sheet_name] = {'written': 0, 'skipped': 0, 'errors': 0}

        print(f"  [{i}/{len(keys)}] {sheet_name}")

        if sheet_name not in tracker_wb.sheetnames:
            print(f"         ERROR: Sheet '{sheet_name}' not in tracker.")
            total_errors += 1
            company_stats[sheet_name]['errors'] += 1
            continue

        folder_path.mkdir(parents=True, exist_ok=True)
        files = sorted(folder_path.glob('*.xlsm'))
        if not files:
            print("         No files.")
            continue

        ws = tracker_wb[sheet_name]
        next_row = get_next_empty_row(ws)
        next_sn  = get_next_sn(ws, next_row)
        print(f"         Next row: {next_row}  Next SN: {next_sn}  Files: {len(files)}")

        for j, fpath in enumerate(files, 1):
            try:
                patient = read_patient(fpath)

                if not patient['iqama']:
                    print(f"         SKIP {fpath.name} - Iqama empty")
                    total_skipped += 1
                    company_stats[sheet_name]['skipped'] += 1
                    continue

                if not patient['status']:
                    print(f"         WARN {fpath.name} - no status checkmark")

                if iqama_exists(ws, patient['iqama']):
                    print(f"         WARN Iqama {patient['iqama']} already exists - adding anyway")

                if args.dry_run:
                    abn = [c for c, v in patient['tests'].items() if v == 'ABNORMAL']
                    print(f"         DRY  row={next_row} {patient['name']} | {patient['iqama']} | status={patient['status']} | abn={abn or 'none'}")
                    total_written += 1
                    company_stats[sheet_name]['written'] += 1
                else:
                    write_row(ws, next_row, next_sn, patient)
                    print(f"         OK   row={next_row} SN={next_sn}  {patient['name']}  ({patient['iqama']})  status={patient['status']}")
                    total_written += 1
                    company_stats[sheet_name]['written'] += 1

                    if not args.no_archive:
                        try:
                            archive_file(fpath, archive_dir / folder_name)
                        except Exception as e:
                            print(f"         WARN archive failed: {e}")

                    next_row += 1
                    next_sn  += 1

            except Exception as e:
                print(f"         ERROR {fpath.name}: {e}")
                total_errors += 1
                company_stats[sheet_name]['errors'] += 1

    # Save tracker
    if not args.dry_run:
        print()
        print("  Saving tracker...")
        try:
            before_size = tracker_path.stat().st_size
            before_mtime = tracker_path.stat().st_mtime
            tracker_wb.save(tracker_path)
            after = tracker_path.stat()
            print(f"  Tracker saved.  ({before_size:,} -> {after.st_size:,} bytes)")
        except Exception as e:
            print(f"  ERROR: Save failed: {e}")
            total_errors += 1
    else:
        print("  DRY-RUN: tracker not modified.")

    tracker_wb.close()

    elapsed = datetime.now() - start_time
    elapsed_str = str(elapsed).split('.')[0]

    print()
    print('  ============================================================')
    print('                          S U M M A R Y                       ')
    print('  ============================================================')
    print()
    mode = 'DRY-RUN' if args.dry_run else 'LIVE (tracker updated)'
    print(f"   Mode:                  {mode}")
    print(f"   Companies scanned:     {len(keys)}")
    print(f"   Patient files written: {total_written}")
    print(f"   Files skipped:         {total_skipped}")
    print(f"   Errors:                {total_errors}")
    print(f"   Time elapsed:          {elapsed_str}")
    print()

    active = {s: v for s, v in company_stats.items()
              if v['written'] or v['skipped'] or v['errors']}
    if active:
        print(f"   {'Company':<22} {'Written':>8} {'Skipped':>8} {'Errors':>8}")
        print(f"   {'-'*22} {'-'*8} {'-'*8} {'-'*8}")
        for sheet, v in active.items():
            print(f"   {sheet:<22} {v['written']:>8} {v['skipped']:>8} {v['errors']:>8}")
        print()

    print(f"   Tracker: {tracker_path}")
    print('  ============================================================')
    if total_errors:
        print('               FINISHED WITH ERRORS / WARNINGS               ')
    else:
        print('                       FINISHED SUCCESSFULLY                  ')
    print('  ============================================================')
    print()

    sys.exit(1 if total_errors else 0)

if __name__ == '__main__':
    main()
